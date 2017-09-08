#!/usr/bin/env python
# encoding: utf-8
# ===============================================================================
#
#         FILE:
#
#        USAGE:
#
#  DESCRIPTION:
#
#      OPTIONS:  ---
# REQUIREMENTS:  ---
#         BUGS:  ---
#        NOTES:  ---
#       AUTHOR:  YOUR NAME (),
#      COMPANY:
#      VERSION:  1.0
#      CREATED:
#     REVISION:  ---
# ===============================================================================

from datetime import datetime
from sys import argv, exit, version_info
from getpass import getpass
import pprint
import getopt
import argparse
# print(version_info)
if version_info < (3, 0):
    raise RuntimeError('At least Python 3.0 is required')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Color, Border, colors, PatternFill, Border, Side, Alignment, Protection
except ImportError:
    print("No openpyxl find, Please pip install openpyxl")
    exit(1)
try:
    import pymysql as mysql
except ImportError:
    print("No pymysql find, Please pip install pymysql")
    exit(1)
try:
    import lib.BashColor as Color
except ImportError:
    class BashColor:
        def __getattr__(self, item):
            return print
    Color = BashColor()

mysql_timeout = 60
mysql_charset = "UTF8"
#单元格内的字体大小
font_size = 12
#单元格最大的宽度
cell_max_width = 30


lang = {
    "zhCN": ["字段", "描述", "是否为空", "默认值", "数据类型", "字符长度", "数字长度", "小数位数"],
    "en": ["Name", "Comment", "Is Nullable","column_default" , "Type", "Character Maximum Length", "Numeric Precision", "Numeric Scale"]
}

class ArgvError(Exception):
    def __init__(self, value):
        self.value = value

    def __str__(self):
        print()
        usage()
        print()
        return repr(self.value)


def get_tmp_file_name():
    return datetime.now().strftime("_%Y-%m-%d_%H%M%S")


def error_exit(info):
    error_msg(info)
    print()
    usage()
    exit(1)


def error_msg(info):
    Color.red('[Error]: ' + str(info))


def get_opts():
    opt = dict()
    # Get opt
    try:
        # print(str(argv))
        opts, args = getopt.getopt(argv[1:], 'h:P:u:p:d:l:o:', ['host=', 'port=', 'user=', 'pass=', 'db=', 'lang=', 'output='])
        #print("debug " + str(opts))
        for key, value in opts:
            if key in ['-h', '--host']:
                opt['host'] = value
            if key in ['-P', '--port']:
                opt['port'] = value
            if key in ['-u', '--user']:
                opt['user'] = value
            if key in ['-p', '--pass']:
                opt['pass'] = value
            if key in ['-d', '--db']:
                opt['db'] = value
            if key in ['-l', '--lang']:
                opt['lang'] = value
            if key in ['-o', '--output']:
                opt['output'] = value
    except getopt.GetoptError as e:
        error_msg(e.msg)
        exit(1)
        # raise ArgvError("Argv error")
    # Check opt
    if 'host' not in opt.keys():
        opt['host'] = input("MySQL Host: ")
    if 'port' not in opt.keys():
        opt['port'] = 3306
        Color.yellow("Port not define, use default, %s" % opt['port'])
    if 'user' not in opt.keys():
        opt['user'] = input("Username: ")
    if 'pass' not in opt.keys():
        opt['pass'] = getpass("Password: ")
    if 'db' not in opt.keys():
        opt['db'] = input("Database: ")
    if 'lang' not in opt.keys():
        opt['lang'] = 'zhCN'
        Color.yellow("Lang not define, use default, %s" % opt['lang'] )
    if 'output' not in opt.keys():
        opt['output'] = 'db_schema_' + opt['db'] + get_tmp_file_name() + '.xlsx'
        Color.yellow("Output not define, save to default, %s" % opt['output'] )
    return opt


def usage():
    first_line = "Usage: %s [ OPTIONS ]" % argv[0]
    break_line = "=" * len(first_line)
    print(break_line)
    print(first_line)
    print("""
Options:
    -h [--host=]  :  MySQL Host
    -u [--user=]  :  MySQL User, Need Read Information_schema
    -u [--pass=]  :  User password
    -d [--db=]    :  Which db you need dump
    -l [--lang=]  :  Column language[todo]
    -o [--output=]:  file to save""")
    #print(break_line)
    print()


def arg_parse():
    parser = argparse.ArgumentParser()
    #https://docs.python.org/3.6/howto/argparse.html#id1
    #todo


def connect_db(opt):
    conn = None
    try:
        conn = mysql.connect(
            host=opt['host'],
            port=opt['port'],
            user=opt['user'],
            password=opt['pass'],
            connect_timeout=mysql_timeout,
            charset=mysql_charset,
            cursorclass=mysql.cursors.DictCursor
        )
    except Exception as e:
        error_msg("DB Error, Msg: %s" % e)
        exit(1)
    return conn


def get_all_table(conn, db):
    result = []
    try:
        with conn.cursor() as cursor:
            sql = "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_SCHEMA=%s"
            cursor.execute(sql, db)
            result = [item['TABLE_NAME'] for item in cursor.fetchall()]
    except Exception as e:
        error_msg("DB Error, Msg: %s" % e)
        exit(1)
    return result
    # return tables


def get_table_info(conn, db, table_name):
    result = {}
    try:
        with conn.cursor() as cursor:
            sql = "select column_name, column_comment, is_nullable, column_default, column_type, character_maximum_length, numeric_precision, numeric_scale from information_schema.columns where table_schema = %s and table_name = %s"
            cursor.execute(sql, (db, table_name))
            result = cursor.fetchall()
    except Exception as e:
        error_msg("DB Error, Msg: %s" % e)
        exit(1)
    return result


def as_text(value):
    if value is None:
        return ""
    # 考虑到中文，转化成byte
    return str(value).encode()

def main():
    # print(str(argv))
    opt = get_opts()
    conn = connect_db(opt)
    tables = {}
    # set border
    left_alignment = Alignment(horizontal='left',
                          vertical='bottom',
                          text_rotation=0,
                          wrap_text=False,
                          shrink_to_fit=True,
                          indent=0)
    right_alignment = Alignment(horizontal='right',
                          vertical='bottom',
                          text_rotation=0,
                          wrap_text=True,
                          shrink_to_fit=True,
                          indent=0)
    center_alignment=Alignment(horizontal='center',
                     vertical='bottom',
                     text_rotation=0,
                     wrap_text=False,
                     shrink_to_fit=False,
                     indent=0)
    bold_font = Font(name="Tahoma", bold=True, size=font_size)
    font = Font(name="Tahoma", size=font_size)

    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="ff0000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    fill_name = PatternFill("solid", fgColor="95B3D7")
    fill_title = PatternFill("solid", fgColor="D9D9D9")

    wb = Workbook()
    # sheet = wb.create_sheet(index=0, title=opt['db'])
    sheet = wb.active
    row = 0
    for index, table_name in enumerate(get_all_table(conn, opt['db'])):
        tables[table_name] = table_info = get_table_info(conn, opt['db'], table_name)
        row = row + 1
        # 这是标题，合并单元格
        # color 95B3D7
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        sheet.cell(row=row, column=1).value = table_name
        sheet.cell(row=row, column=1).font = font
        sheet.cell(row=row, column=1).fill = fill_name
        sheet.cell(row=row, column=1).alignment = center_alignment
        for x in range(1, 9):
            sheet.cell(row=row, column=x).border = border

        row = row + 1
        title = lang[opt['lang']]
        for col, val in enumerate(title, start=1):
            # color D9D9D9
            sheet.cell(row=row, column=col).value = val
            sheet.cell(row=row, column=col).font = bold_font
            # sheet.cell(row=row, column=col).border = border
            sheet.cell(row=row, column=col).fill = fill_title
            sheet.cell(row=row, column=col).border = border

        for line in table_info:
            # ["字段", "描述", "是否必填", "默认值", "数据类型", "字符长度", "数字长度", "小数位数"],
            row = row + 1
            col = 1
            sheet.cell(row=row, column=col).value = line['column_name']
            sheet.cell(row=row, column=col).font = font
            sheet.cell(row=row, column=col).border = border
            col = col + 1
            sheet.cell(row=row, column=col).value = line['column_comment']
            sheet.cell(row=row, column=col).font = font
            sheet.cell(row=row, column=col).border = border
            col = col + 1
            sheet.cell(row=row, column=col).value = line['is_nullable']
            sheet.cell(row=row, column=col).font = font
            sheet.cell(row=row, column=col).border = border
            col = col + 1
            sheet.cell(row=row, column=col).value = line['column_default']
            sheet.cell(row=row, column=col).font = font
            sheet.cell(row=row, column=col).border = border
            col = col + 1
            sheet.cell(row=row, column=col).value = line['column_type']
            sheet.cell(row=row, column=col).font = font
            sheet.cell(row=row, column=col).border = border
            col = col + 1
            sheet.cell(row=row, column=col).value = line['character_maximum_length']
            sheet.cell(row=row, column=col).font = font
            sheet.cell(row=row, column=col).border = border
            col = col + 1
            sheet.cell(row=row, column=col).value = line['numeric_precision']
            sheet.cell(row=row, column=col).font = font
            sheet.cell(row=row, column=col).border = border
            col = col + 1
            sheet.cell(row=row, column=col).value = line['numeric_scale']
            sheet.cell(row=row, column=col).font = font
            sheet.cell(row=row, column=col).border = border

        row = row + 1
        sheet.cell(row=row, column=1).value = ""
    # print("done: %s" % str(opt))
    # pprint.pprint(tables)

    # 调整单元格的宽度
    for column_cells in sheet.columns:
        length = max(len(as_text(cell.value)) for cell in column_cells)
        # print("%s: %s" %(column_cells, length))
        # print()
        if length > cell_max_width:
            length = cell_max_width
        sheet.column_dimensions[column_cells[0].column].width = length
    # save to file
    print("save to %s" % opt['output'])
    # start to save xlsx
    wb.save(opt['output'])


if __name__ == '__main__':
    main()

