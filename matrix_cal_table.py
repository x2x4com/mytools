#!/usr/bin/env python
# encoding: utf-8
# ===============================================================================
#
#         FILE:
#
#        USAGE:
#
#  DESCRIPTION:
#
#      OPTIONS:  ---
# REQUIREMENTS:  ---
#         BUGS:  ---
#        NOTES:  ---
#       AUTHOR:  YOUR NAME (),
#      COMPANY:
#      VERSION:  1.0
#      CREATED:
#     REVISION:  ---
# ===============================================================================

from datetime import datetime
from sys import argv, exit, version_info
import argparse
import random
import re
# print(version_info)
if version_info < (3, 0):
    raise RuntimeError('At least Python 3.0 is required')

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Color, Border, colors, PatternFill, Border, Side, Alignment, Protection
except ImportError:
    print("No openpyxl find, Please pip install openpyxl")
    exit(1)

try:
    import lib.BashColor as Color
except ImportError:
    class BashColor:
        def __getattr__(self, item):
            return print
    Color = BashColor()

#单元格内的字体大小
font_size = 18
#单元格最大的宽度
cell_max_width = 10
cell_max_height = 50
re_num_ranger = re.compile(r'^(\d+)-(\d+)$')
default_start = 1
default_stop = 20



def get_tmp_file_name():
    return datetime.now().strftime("_%Y-%m-%d_%H%M%S")


def arg_parse():
    parser = argparse.ArgumentParser(description="create calculate table")
    # https://docs.python.org/3.6/howto/argparse.html#id1
    # group = parser.add_mutually_exclusive_group()
    # group.add_argument("-v", "--verbose", action="store_true")
    # group.add_argument("-q", "--quiet", action="store_true")
    # parser.add_argument("-v", "--verbosity", type=int, choices=[0, 1, 2], help="increase output verbosity")
    #
    # action : store_true 当参数出现时候标记为true
    # action : count 统计参数出现的次数  -v v=1  -vv v=2 -vvv v=3
    # choices : 选择
    # default : 默认值
    parser.add_argument("-a", "--amount", help="col/row amount", type=int)
    parser.add_argument("-m", "--max", help="col + row < max", type=int)
    parser.add_argument("-c", "--column", help="column number ranger, example: 1-10")
    parser.add_argument("-r", "--row", help="row number ranger, example: 1-10")
    parser.add_argument("-o", "--output", help="output excel file name, default is random")
    # parser.add_argument("--hello", help="col/row amount")
    args = parser.parse_args()
    # print(args)
    # print(args.a)
    # parser.exit(1,"some words")
    # parser.error("some error")
    if args.amount is None:
        parser.print_help()
        parser.exit("amount not defined")

    if args.column is None or not get_number(args.column, True):
        parser.print_help()
        parser.exit("column not defined")

    if args.row is None or not get_number(args.row, True):
        parser.print_help()
        parser.exit("row not defined")

    if args.max is None:
        parser.print_help()
        parser.exit("max not defined")

    if args.output is None:
        args.output = "output" + get_tmp_file_name() + ".xlsx"

    return args


def get_number(need, test=False):
    # \d+-\d+
    # print("word: %s" % need)
    result = re_num_ranger.match(need)
    # print("re result: %s" % result)
    if test:
        if result:
            return True
        else:
            return False
    else:
        if result:
            start = int(result.group(1))
            stop = int(result.group(2))
            return random.randrange(start, stop)
        else:
            return random.randrange(default_start,default_stop)

def run(args):
    left_alignment = Alignment(horizontal='left',
                               vertical='center',
                               text_rotation=0,
                               wrap_text=False,
                               shrink_to_fit=True,
                               indent=0)
    right_alignment = Alignment(horizontal='right',
                                vertical='center',
                                text_rotation=0,
                                wrap_text=True,
                                shrink_to_fit=True,
                                indent=0)
    center_alignment = Alignment(horizontal='center',
                                 vertical='center',
                                 text_rotation=0,
                                 wrap_text=False,
                                 shrink_to_fit=False,
                                 indent=0)
    bold_font = Font(name="Tahoma", bold=True, size=font_size)
    font = Font(name="Tahoma", size=font_size)

    thin = Side(border_style="thin", color="000000")
    double = Side(border_style="double", color="ff0000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    fill_name = PatternFill("solid", fgColor="95B3D7")
    fill_title = PatternFill("solid", fgColor="D9D9D9")
    wb = Workbook()
    sheet = wb.active

    row = 1
    row_1_data = dict()
    col_1_data = dict()
    while row <= (args.amount + 1):
        # print("第%s行, 共计%s行" % (row, (args.amount + 1)))
        if row == 1:
            # 第一行， 第一列是空的，第二列开始是数字
            col = 1
            while col <= (args.amount + 1):
                # print("第%s行, 第%s列, 共%s列" % (row, col, (args.amount + 1)))
                if col == 1:
                    # 第一行，第一列空
                    sheet.cell(row=row, column=col).value = ""
                    sheet.cell(row=row, column=col).font = font
                    sheet.cell(row=row, column=col).border = border
                    col += 1
                    continue
                num = get_number(args.column)
                row_1_data[col] = num
                sheet.cell(row=row, column=col).value = num
                sheet.cell(row=row, column=col).font = font
                sheet.cell(row=row, column=col).border = border
                sheet.cell(row=row, column=col).alignment = center_alignment
                col += 1
        if row > 1:
            col = 1
            while col <= (args.amount + 1):
                # print("第%s行, 第%s列, 共%s列" % (row, col, (args.amount + 1)))
                if col == 1:
                    not_ok = True
                    num = 0
                    while not_ok:
                        num = get_number(args.row)
                        for row_data in row_1_data:
                            if (num + row_data) > args.max:
                                not_ok = True
                            else:
                                not_ok = False

                    sheet.cell(row=row, column=col).value = num
                    sheet.cell(row=row, column=col).font = font
                    sheet.cell(row=row, column=col).border = border
                    sheet.cell(row=row, column=col).alignment = center_alignment
                    col += 1
                    continue
                if col > 1:
                    sheet.cell(row=row, column=col).value = ""
                    sheet.cell(row=row, column=col).font = font
                    sheet.cell(row=row, column=col).border = border
                    col += 1
        sheet.row_dimensions[row].height = cell_max_height
        row += 1

    for column_cells in sheet.columns:
        sheet.column_dimensions[column_cells[0].column].width = cell_max_width
    print("save to %s" % args.output)
    wb.save(args.output)


def main():
    args = arg_parse()
    run(args)


if __name__ == '__main__':
    main()
